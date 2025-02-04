import openpyxl


class HomepageData:
    test_Homepage_Data=[{"firstnane":"Rahul", "lastnane":"Aggarwal" ,"email":"rk@gmail.com"}, {"firstnane":"Nikita", "lastnane":"Gupta", "email":"ng@gmail.com"}]


    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\surya\\OneDrive\\Desktop\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]